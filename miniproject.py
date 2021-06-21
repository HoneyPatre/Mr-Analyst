import tkinter as tk
from tkinter import font as tkfont
import matplotlib.pyplot as plt
from openpyxl import *
from tkinter import *
from PIL import Image,ImageTk
from tkinter.filedialog import askopenfilename
import pandas as pd

class Application(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.title_font = tkfont.Font(family='Helvetica', size=40, weight="bold", slant="italic")
        self.other_font = tkfont.Font(family='Times', size=30)
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (MainMenu, MyStudents, Marks):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("MainMenu")

    def show_frame(self, page_name):
        # Show a frame for the given page name
        frame = self.frames[page_name]
        frame.tkraise()


class MainMenu(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        logo = tk.PhotoImage(file="images/main.png")
        BGlabel = tk.Label(self, image=logo)
        BGlabel.image = logo
        BGlabel.place(x=0, y=0, width=1920, height=1080)

        label = tk.Label(self, text="Mr. Analyst", font=controller.title_font)
        label.pack(side="top", pady=10)
        button1 = tk.Button(self, text="My Students", font=controller.other_font ,command=lambda: controller.show_frame("MyStudents"))
        button2 = tk.Button(self, text="Marks Evaluation", font=controller.other_font ,command=lambda: controller.show_frame("Marks"))
        button3 = tk.Button(self, text="Attendance", font=controller.other_font , command=lambda: controller.show_frame("Attendance"))
        button1.pack(side=TOP,pady=10,padx=10)
        button2.pack(side=TOP,pady=10,padx=10)
        button3.pack(side=TOP,pady=10,padx=10)

class MyStudents(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        wb = load_workbook('data/test.xlsx')
        self.sheet = wb.active
        self.controller = controller

        logo = tk.PhotoImage(file="images/main.png")
        BGlabel = tk.Label(self, image=logo)
        BGlabel.image = logo
        BGlabel.place(x=0, y=0, width=1920, height=1080)


        label = tk.Label(self, text="Student Profiles", font=controller.title_font)
        label.pack(side="top", pady=10)
        self.newstudent = tk.Button(self, text="New Registrations",font=controller.other_font , command=NewRegistration)
        self.newstudent.pack(side=TOP,pady=10,padx=10)

        button = tk.Button(self, text="Main Menu", font=controller.other_font ,command=lambda: controller.show_frame("MainMenu"))
        button.pack(side=TOP,pady=10,padx=10)


class NewRegistration:
    def __init__(self):
        self.wb = load_workbook('data/test.xlsx')
        self.sheet = self.wb.active
        root = tk.Toplevel()

        title_font = tkfont.Font(family='Helvetica', size=40, weight="bold", slant="italic")
        other_font = tkfont.Font(family='Times', size=30)

        img = Image.open("images/registration_form.png")
        logo = ImageTk.PhotoImage(img)
        BGlabel = tk.Label(root,image=logo)
        BGlabel.place(x=0, y=0, width=1920, height=1080)

        root.title("Registration Form")
        root.geometry("1920x1080")
        self.excel()

        heading = Label(root, text="Form",font=title_font)
        name = Label(root, text="Name",font=other_font)
        course = Label(root, text="SSC",font=other_font)
        sem = Label(root, text="HSC",font=other_font)
        form_no = Label(root, text="Roll No.",font=other_font)
        contact_no = Label(root, text="Contact No.",font=other_font)
        email_id = Label(root, text="Email id",font=other_font)
        address = Label(root, text="Address",font=other_font)

        heading.grid(row=0, column=1)

        self.name_field = Entry(root,font=other_font)
        self.course_field = Entry(root,font=other_font)
        self.sem_field = Entry(root,font=other_font)
        self.form_no_field = Entry(root,font=other_font)
        self.contact_no_field = Entry(root,font=other_font)
        self.email_id_field = Entry(root,font=other_font)
        self.address_field = Entry(root,font=other_font)

        self.name_field.bind("<Return>", self.focus1)
        self.course_field.bind("<Return>", self.focus2)
        self.sem_field.bind("<Return>", self.focus3)
        self.form_no_field.bind("<Return>", self.focus4)
        self.contact_no_field.bind("<Return>", self.focus5)
        self.email_id_field.bind("<Return>", self.focus6)

        form_no.grid(row=1, column=0)
        self.form_no_field.grid(row=1, column=1, ipadx="100")
        name.grid(row=2, column=0)
        self.name_field.grid(row=2, column=1, ipadx="100")
        contact_no.grid(row=3, column=0)
        self.contact_no_field.grid(row=3, column=1, ipadx="100")
        email_id.grid(row=4, column=0)
        self.email_id_field.grid(row=4, column=1, ipadx="100")
        address.grid(row=5, column=0)
        self.address_field.grid(row=5, column=1, ipadx="100")
        course.grid(row=6, column=0)
        self.course_field.grid(row=6, column=1, ipadx="100")
        sem.grid(row=7, column=0)
        self.sem_field.grid(row=7, column=1, ipadx="100")

        self.excel()
        submit = Button(root, text="Submit", fg="Black", bg="green", command=self.insert,font=other_font)
        submit.grid(row=8, column=1)

        root.grid()
        root.mainloop()

    def excel(self):

        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 10
        self.sheet.column_dimensions['C'].width = 10
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 40
        self.sheet.column_dimensions['G'].width = 50

        self.sheet.cell(row=1, column=1).value = "Name"
        self.sheet.cell(row=1, column=2).value = "Course"
        self.sheet.cell(row=1, column=3).value = "Semester"
        self.sheet.cell(row=1, column=4).value = "Form Number"
        self.sheet.cell(row=1, column=5).value = "Contact Number"
        self.sheet.cell(row=1, column=6).value = "Email id"
        self.sheet.cell(row=1, column=7).value = "Address"

    def focus1(self, event):
        self.course_field.focus_set()

    def focus2(self, event):
        self.sem_field.focus_set()

    def focus3(self, event):
        self.form_no_field.focus_set()

    def focus4(self, event):
        self.contact_no_field.focus_set()

    def focus5(self, event):
        self.email_id_field.focus_set()

    def focus6(self, event):
        self.address_field.focus_set()

    def clear(self):
        self.name_field.delete(0, END)
        self.course_field.delete(0, END)
        self.sem_field.delete(0, END)
        self.form_no_field.delete(0, END)
        self.contact_no_field.delete(0, END)
        self.email_id_field.delete(0, END)
        self.address_field.delete(0, END)

    def insert(self):
        if (self.name_field.get() == "" and
                self.course_field.get() == "" and
                self.sem_field.get() == "" and
                self.form_no_field.get() == "" and
                self.contact_no_field.get() == "" and
                self.email_id_field.get() == "" and
                self.address_field.get() == ""):

            print("empty input")

        else:

            current_row = self.sheet.max_row
            self.sheet.cell(row=current_row + 1, column=1).value = self.form_no_field.get()
            self.sheet.cell(row=current_row + 1, column=2).value = self.name_field.get()
            self.sheet.cell(row=current_row + 1, column=3).value = self.contact_no_field.get()
            self.sheet.cell(row=current_row + 1, column=4).value = self.email_id_field.get()
            self.sheet.cell(row=current_row + 1, column=5).value = self.address_field.get()
            self.sheet.cell(row=current_row + 1, column=6).value = self.course_field.get()
            self.sheet.cell(row=current_row + 1, column=7).value = self.sem_field.get()

            self.wb.save('data/test.xlsx')
            self.name_field.focus_set()
            self.clear()


class Marks(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        logo = tk.PhotoImage(file="images/marks.png")
        BGlabel = tk.Label(self, image=logo)
        BGlabel.image = logo
        BGlabel.place(x=0, y=0, width=1920, height=1080)

        label = tk.Label(self, text="Marks Evaluation", font=controller.title_font)
        label.pack(side="top", pady=10)
        self.UT_marks = tk.Button(self, text="UT-1 Marks", command=self.marksUT1,font=controller.other_font)
        self.UT_marks.pack(side=TOP,pady=10,padx=10)
        self.UT_marks = tk.Button(self, text="UT-2 Marks", command=self.marksUT2, font=controller.other_font)
        self.UT_marks.pack(side=TOP, pady=10, padx=10)
        self.UT_marks = tk.Button(self, text="In-Sem Marks", command=self.marksInSem, font=controller.other_font)
        self.UT_marks.pack(side=TOP, pady=10, padx=10)
        self.UT_marks = tk.Button(self, text="End-Sem Marks", command=self.marksEndSem, font=controller.other_font)
        self.UT_marks.pack(side=TOP, pady=10, padx=10)
        button = tk.Button(self, text="Main Menu", command=lambda: controller.show_frame("MainMenu"),font=controller.other_font)
        button.pack(side=TOP,pady=10,padx=10)

    def marksUT1(self):
        filename = askopenfilename()
        print(filename)
        data_xls = pd.read_excel(filename, 'Sheet1', index_col=None)
        data_xls.to_csv('data/Unittest1.csv', encoding='utf-8')

        df = pd.read_csv("data/Unittest1.csv", sep=",").set_index("Roll_no")
        d = dict(zip(df.index, df.values.tolist()))
        print(d)
        df.set_index('student')[['sub1', 'sub2', 'sub3']].plot.bar()
        plt.show()

    def marksUT2(self):
        filename = askopenfilename()
        print(filename)
        data_xls = pd.read_excel(filename, 'Sheet1', index_col=None)
        data_xls.to_csv('data/Unittest2.csv', encoding='utf-8')

        df = pd.read_csv("data/Unittest2.csv", sep=",").set_index("RollNo")
        d = dict(zip(df.index, df.values.tolist()))
        print(d)
        df.set_index('Name')[['CN', 'DBMS', 'ISEE','SEPM','TOC']].plot.bar()
        plt.show()

    def marksInSem(self):
        filename = askopenfilename()
        print(filename)
        data_xls = pd.read_excel(filename, 'Sheet1', index_col=None)
        data_xls.to_csv('data/InSem.csv', encoding='utf-8')

        df = pd.read_csv("data/InSem.csv", sep=",").set_index("RollNo")
        d = dict(zip(df.index, df.values.tolist()))
        print(d)
        df.set_index('Name')[['CN', 'DBMS', 'ISEE','SEPM','TOC']].plot.bar()
        plt.show()

    def marksEndSem(self):
        filename = askopenfilename()
        print(filename)
        data_xls = pd.read_excel(filename, 'Sheet1', index_col=None)
        data_xls.to_csv('data/Endsem.csv', encoding='utf-8')

        df = pd.read_csv("data/Endsem.csv", sep=",").set_index("RollNo")
        d = dict(zip(df.index, df.values.tolist()))
        print(d)
        df.set_index('Name')[['CN', 'DBMS', 'ISEE','SEPM','TOC']].plot.bar()
        plt.show()


class Attendance(tk.Frame):
    pass

if __name__ == "__main__":
    app = Application()
    app.geometry("1980x1020")
    app.mainloop()
